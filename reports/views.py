from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import date, timedelta
import json
import io
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .models import Report, ReportTemplate, ReportShare
from .forms import (
    ReportForm, PatientsReportForm, AppointmentsReportForm, 
    FinancialReportForm, AnalyticsReportForm, ReportTemplateForm,
    ReportShareForm, CustomReportForm
)
from patients.models import Patient
from appointments.models import Appointment
from medical_records.models import MedicalRecord

@login_required  
def index(request):
    """Dashboard principal de reportes"""
    # Obtener reportes recientes del usuario
    recent_reports = Report.objects.filter(
        created_by=request.user
    ).order_by('-created_at')[:10]
    
    # Estadísticas de reportes
    stats = {
        'total_reports': Report.objects.filter(created_by=request.user).count(),
        'completed_reports': Report.objects.filter(
            created_by=request.user, 
            status='completed'
        ).count(),
        'scheduled_reports': Report.objects.filter(
            created_by=request.user, 
            is_scheduled=True
        ).count(),
        'shared_reports': ReportShare.objects.filter(
            shared_with=request.user
        ).count(),
    }
    
    # Plantillas disponibles
    templates = ReportTemplate.objects.filter(
        Q(created_by=request.user) | Q(is_public=True),
        is_active=True
    )[:5]
    
    context = {
        'title': 'Reportes',
        'recent_reports': recent_reports,
        'stats': stats,
        'templates': templates,
    }
    
    return render(request, 'reports/index.html', context)

@login_required
def report_list(request):
    """Lista todos los reportes del usuario"""
    reports_list = Report.objects.filter(created_by=request.user).order_by('-created_at')
    
    # Filtros
    status_filter = request.GET.get('status')
    type_filter = request.GET.get('type')
    
    if status_filter:
        reports_list = reports_list.filter(status=status_filter)
    if type_filter:
        reports_list = reports_list.filter(report_type=type_filter)
    
    # Paginación
    paginator = Paginator(reports_list, 15)
    page_number = request.GET.get('page')
    reports = paginator.get_page(page_number)
    
    context = {
        'title': 'Mis Reportes',
        'reports': reports,
        'status_filter': status_filter,
        'type_filter': type_filter,
    }
    
    return render(request, 'reports/list.html', context)

@login_required
def report_detail(request, pk):
    """Detalle de un reporte específico"""
    report = get_object_or_404(Report, pk=pk)
    
    # Verificar permisos
    if report.created_by != request.user:
        # Verificar si está compartido con el usuario
        try:
            share = ReportShare.objects.get(report=report, shared_with=request.user)
            if share.is_expired:
                raise Http404("El acceso a este reporte ha expirado")
        except ReportShare.DoesNotExist:
            raise Http404("No tienes permisos para ver este reporte")
    
    context = {
        'title': f'Reporte: {report.title}',
        'report': report,
    }
    
    return render(request, 'reports/detail.html', context)

@login_required
def patients_report(request):
    """Generar reporte de pacientes"""
    form = PatientsReportForm(request.GET or None)
    patients_data = None
    
    if form.is_valid():
        # Construir consulta base
        queryset = Patient.objects.all()
        
        # Aplicar filtros
        if form.cleaned_data.get('date_from'):
            queryset = queryset.filter(registration_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            queryset = queryset.filter(registration_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('gender'):
            queryset = queryset.filter(gender=form.cleaned_data['gender'])
        if form.cleaned_data.get('city'):
            queryset = queryset.filter(city__icontains=form.cleaned_data['city'])
        if form.cleaned_data.get('active_only'):
            queryset = queryset.filter(is_active=True)
        
        # Filtro por rango de edad
        age_range = form.cleaned_data.get('age_range')
        if age_range:
            today = date.today()
            if age_range == '0-18':
                min_birth = today - timedelta(days=18*365)
                queryset = queryset.filter(birth_date__gte=min_birth)
            elif age_range == '19-30':
                min_birth = today - timedelta(days=30*365)
                max_birth = today - timedelta(days=19*365)
                queryset = queryset.filter(birth_date__range=[min_birth, max_birth])
            elif age_range == '31-50':
                min_birth = today - timedelta(days=50*365)
                max_birth = today - timedelta(days=31*365)
                queryset = queryset.filter(birth_date__range=[min_birth, max_birth])
            elif age_range == '51-70':
                min_birth = today - timedelta(days=70*365)
                max_birth = today - timedelta(days=51*365)
                queryset = queryset.filter(birth_date__range=[min_birth, max_birth])
            elif age_range == '70+':
                max_birth = today - timedelta(days=70*365)
                queryset = queryset.filter(birth_date__lte=max_birth)
        
        # Calculate average age manually to avoid SQLite aggregation issues
        all_patients = queryset.all()
        total_age = sum(patient.get_age() for patient in all_patients)
        avg_age = total_age / len(all_patients) if len(all_patients) > 0 else 0
        
        patients_data = {
            'patients': queryset.order_by('-registration_date')[:100],  # Limitar para preview
            'total_count': queryset.count(),
            'stats': {
                'by_gender': queryset.values('gender').annotate(count=Count('id')),
                'by_city': queryset.values('city').annotate(count=Count('id')).order_by('-count')[:10],
                'avg_age': round(avg_age, 1),
            }
        }
    
    context = {
        'title': 'Reporte de Pacientes',
        'form': form,
        'patients_data': patients_data,
    }
    
    return render(request, 'reports/patients.html', context)

@login_required
def appointments_report(request):
    """Generar reporte de citas"""
    form = AppointmentsReportForm(request.GET or None)
    appointments_data = None
    
    if form.is_valid():
        # Construir consulta base
        queryset = Appointment.objects.select_related('patient', 'doctor')
        
        # Aplicar filtros
        if form.cleaned_data.get('date_from'):
            queryset = queryset.filter(appointment_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            queryset = queryset.filter(appointment_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('status'):
            queryset = queryset.filter(status=form.cleaned_data['status'])
        if form.cleaned_data.get('doctor'):
            queryset = queryset.filter(doctor=form.cleaned_data['doctor'])
        
        appointments_data = {
            'appointments': queryset.order_by('-appointment_date')[:100],
            'total_count': queryset.count(),
            'stats': {
                'by_status': queryset.values('status').annotate(count=Count('id')),
                'by_doctor': queryset.values('doctor__first_name', 'doctor__last_name').annotate(count=Count('id')).order_by('-count')[:10],
                'by_day': queryset.values('appointment_date__date').annotate(count=Count('id')).order_by('appointment_date__date')[:30],
            }
        }
    
    context = {
        'title': 'Reporte de Citas',
        'form': form,
        'appointments_data': appointments_data,
    }
    
    return render(request, 'reports/appointments.html', context)

@login_required
def financial_report(request):
    """Generar reporte financiero"""
    form = FinancialReportForm(request.GET or None)
    financial_data = None
    
    if form.is_valid():
        # Aquí iría la lógica para el reporte financiero
        # Por ahora, datos de ejemplo
        financial_data = {
            'total_revenue': 15000.00,
            'total_expenses': 8000.00,
            'net_profit': 7000.00,
            'pending_payments': 2500.00,
            'monthly_trends': [
                {'month': 'Enero', 'revenue': 5000, 'expenses': 2500},
                {'month': 'Febrero', 'revenue': 5500, 'expenses': 2800},
                {'month': 'Marzo', 'revenue': 4500, 'expenses': 2700},
            ]
        }
    
    context = {
        'title': 'Reporte Financiero',
        'form': form,
        'financial_data': financial_data,
    }
    
    return render(request, 'reports/financial.html', context)

@login_required
def analytics_report(request):
    """Generar análisis avanzado"""
    form = AnalyticsReportForm(request.GET or None)
    analytics_data = None
    
    if form.is_valid():
        metrics = form.cleaned_data.get('metrics', [])
        analytics_data = {}
        
        if 'patient_growth' in metrics:
            # Crecimiento de pacientes por mes
            patient_growth = Patient.objects.extra(
                select={'month': "DATE_FORMAT(created_at, '%%Y-%%m')"}
            ).values('month').annotate(count=Count('id')).order_by('month')
            analytics_data['patient_growth'] = list(patient_growth)
        
        if 'appointment_trends' in metrics:
            # Tendencias de citas
            appointment_trends = Appointment.objects.extra(
                select={'week': "WEEK(appointment_date)"}
            ).values('week').annotate(count=Count('id')).order_by('week')
            analytics_data['appointment_trends'] = list(appointment_trends)
        
        # Más métricas...
    
    context = {
        'title': 'Análisis Avanzado',
        'form': form,
        'analytics_data': analytics_data,
    }
    
    return render(request, 'reports/analytics.html', context)

@login_required
def create_report(request):
    """Crear un nuevo reporte"""
    if request.method == 'POST':
        form = ReportForm(request.POST, user=request.user)
        if form.is_valid():
            report = form.save()
            messages.success(request, f'Reporte "{report.title}" creado exitosamente.')
            return redirect('reports:detail', pk=report.pk)
    else:
        form = ReportForm(user=request.user)
    
    context = {
        'title': 'Crear Reporte',
        'form': form,
    }
    
    return render(request, 'reports/create.html', context)

@login_required
def custom_report(request):
    """Crear reporte personalizado"""
    form = CustomReportForm(request.GET or None)
    
    if form.is_valid():
        # Procesar reporte personalizado
        pass
    
    context = {
        'title': 'Reporte Personalizado',
        'form': form,
    }
    
    return render(request, 'reports/custom.html', context)

@login_required
def templates_list(request):
    """Lista de plantillas de reportes"""
    templates = ReportTemplate.objects.filter(
        Q(created_by=request.user) | Q(is_public=True),
        is_active=True
    ).order_by('name')
    
    context = {
        'title': 'Plantillas de Reportes',
        'templates': templates,
    }
    
    return render(request, 'reports/templates.html', context)

@login_required
def create_template(request):
    """Crear plantilla de reporte"""
    if request.method == 'POST':
        form = ReportTemplateForm(request.POST, user=request.user)
        if form.is_valid():
            template = form.save()
            messages.success(request, f'Plantilla "{template.name}" creada exitosamente.')
            return redirect('reports:templates')
    else:
        form = ReportTemplateForm(user=request.user)
    
    context = {
        'title': 'Crear Plantilla',
        'form': form,
    }
    
    return render(request, 'reports/create_template.html', context)

@login_required
def share_report(request, pk):
    """Compartir un reporte"""
    report = get_object_or_404(Report, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        form = ReportShareForm(request.POST, user=request.user, report=report)
        if form.is_valid():
            share = form.save()
            messages.success(request, f'Reporte compartido con {share.shared_with.get_full_name()}.')
            return redirect('reports:detail', pk=report.pk)
    else:
        form = ReportShareForm(user=request.user, report=report)
    
    context = {
        'title': f'Compartir: {report.title}',
        'form': form,
        'report': report,
    }
    
    return render(request, 'reports/share.html', context)

@login_required
def download_report(request, pk):
    """Descargar un reporte"""
    report = get_object_or_404(Report, pk=pk)
    
    # Verificar permisos
    can_download = False
    if report.created_by == request.user:
        can_download = True
    else:
        try:
            share = ReportShare.objects.get(
                report=report, 
                shared_with=request.user,
                can_download=True
            )
            if not share.is_expired:
                can_download = True
        except ReportShare.DoesNotExist:
            pass
    
    if not can_download:
        raise Http404("No tienes permisos para descargar este reporte")
    
    if not report.is_ready:
        messages.error(request, "El reporte aún no está listo para descargar.")
        return redirect('reports:detail', pk=pk)
    
    # Generar y servir archivo para descarga
    try:
        from django.http import HttpResponse
        import json
        
        # Generar contenido del reporte
        report_data = report.generate_export_data()
        
        if report.report_format == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}.pdf"'
            # Aquí se podría usar reportlab para generar PDF real
            response.write(b'%PDF-1.4\n%Sample PDF content\n')
        elif report.report_format == 'excel':
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}.xlsx"'
            # Aquí se podría usar openpyxl para generar Excel real
            response.write(b'Sample Excel content')
        else:  # JSON por defecto
            response = HttpResponse(json.dumps(report_data, indent=2), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="reporte_{report.id}.json"'
        
        messages.success(request, "Reporte descargado exitosamente.")
        return response
        
    except Exception as e:
        messages.error(request, f"Error al generar el archivo de descarga: {str(e)}")
        return redirect('reports:detail', pk=pk)

# API Views
@login_required
def api_report_status(request, pk):
    """API para obtener el estado de un reporte"""
    try:
        report = Report.objects.get(pk=pk, created_by=request.user)
        return JsonResponse({
            'status': report.status,
            'progress': 100 if report.status == 'completed' else 50,
            'is_ready': report.is_ready,
            'file_size': report.file_size,
        })
    except Report.DoesNotExist:
        return JsonResponse({'error': 'Reporte no encontrado'}, status=404)

@login_required
def api_delete_report(request, pk):
    """API para eliminar un reporte"""
    if request.method == 'DELETE':
        try:
            report = Report.objects.get(pk=pk, created_by=request.user)
            report.delete()
            return JsonResponse({'success': True})
        except Report.DoesNotExist:
            return JsonResponse({'error': 'Reporte no encontrado'}, status=404)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def get_filtered_patients_queryset(request):
    """Helper function to get filtered patients based on request parameters"""
    form = PatientsReportForm(request.GET or None)
    queryset = Patient.objects.all()
    
    if form.is_valid():
        # Apply filters (same logic as patients_report view)
        if form.cleaned_data.get('date_from'):
            queryset = queryset.filter(registration_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            queryset = queryset.filter(registration_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('gender'):
            queryset = queryset.filter(gender=form.cleaned_data['gender'])
        if form.cleaned_data.get('city'):
            queryset = queryset.filter(city__icontains=form.cleaned_data['city'])
        if form.cleaned_data.get('active_only'):
            queryset = queryset.filter(is_active=True)
            
        # Age range filtering
        age_range = form.cleaned_data.get('age_range')
        if age_range:
            today = date.today()
            if age_range == '0-18':
                min_birth = today - timedelta(days=18*365)
                queryset = queryset.filter(birth_date__gte=min_birth)
            elif age_range == '19-30':
                min_birth = today - timedelta(days=30*365)
                max_birth = today - timedelta(days=19*365)
                queryset = queryset.filter(birth_date__range=[min_birth, max_birth])
            elif age_range == '31-50':
                min_birth = today - timedelta(days=50*365)
                max_birth = today - timedelta(days=31*365)
                queryset = queryset.filter(birth_date__range=[min_birth, max_birth])
            elif age_range == '51-70':
                min_birth = today - timedelta(days=70*365)
                max_birth = today - timedelta(days=51*365)
                queryset = queryset.filter(birth_date__range=[min_birth, max_birth])
            elif age_range == '70+':
                max_birth = today - timedelta(days=70*365)
                queryset = queryset.filter(birth_date__lte=max_birth)
    
    return queryset

@login_required
def export_patients_excel(request):
    """Export patients report to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return export_patients_csv(request)
    
    # Get filtered patients
    queryset = get_filtered_patients_queryset(request)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Pacientes"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws.merge_cells('A1:H1')
    ws['A1'] = "REPORTE DE PACIENTES - TOPICTALES BIOMÉDICA"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment
    
    # Add date
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = header_alignment
    
    # Headers
    headers = [
        'ID Paciente', 'Nombre Completo', 'Género', 'Edad', 
        'Ciudad', 'Teléfono', 'Email', 'Estado'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, patient in enumerate(queryset, 5):
        ws.cell(row=row, column=1, value=patient.patient_id)
        ws.cell(row=row, column=2, value=patient.get_full_name())
        ws.cell(row=row, column=3, value=patient.get_gender_display())
        ws.cell(row=row, column=4, value=patient.get_age())
        ws.cell(row=row, column=5, value=patient.city or "-")
        ws.cell(row=row, column=6, value=patient.phone_number or "-")
        ws.cell(row=row, column=7, value=patient.email or "-")
        ws.cell(row=row, column=8, value="Activo" if patient.is_active else "Inactivo")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_pacientes_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

def export_patients_csv(request):
    """Fallback CSV export for patients"""
    queryset = get_filtered_patients_queryset(request)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reporte_pacientes_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
    
    # Add BOM for proper UTF-8 encoding in Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'ID Paciente', 'Nombre Completo', 'Género', 'Edad',
        'Ciudad', 'Teléfono', 'Email', 'Estado'
    ])
    
    for patient in queryset:
        writer.writerow([
            patient.patient_id,
            patient.get_full_name(),
            patient.get_gender_display(),
            patient.get_age(),
            patient.city or "-",
            patient.phone_number or "-", 
            patient.email or "-",
            "Activo" if patient.is_active else "Inactivo"
        ])
    
    return response

@login_required
def export_patients_pdf(request):
    """Export patients report to PDF format"""
    queryset = get_filtered_patients_queryset(request)
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("REPORTE DE PACIENTES", title_style)
    elements.append(title)
    
    subtitle = Paragraph(
        f"TopicTales Biomédica - Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 20))
    
    # Summary statistics
    total_patients = queryset.count()
    active_patients = queryset.filter(is_active=True).count()
    
    summary_data = [
        ['Total de Pacientes:', str(total_patients)],
        ['Pacientes Activos:', str(active_patients)],
        ['Pacientes Inactivos:', str(total_patients - active_patients)]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Patients table
    data = [['ID', 'Nombre', 'Género', 'Edad', 'Ciudad', 'Estado']]
    
    for patient in queryset[:50]:  # Limit to first 50 for PDF
        data.append([
            patient.patient_id,
            patient.get_full_name()[:25] + '...' if len(patient.get_full_name()) > 25 else patient.get_full_name(),
            patient.get_gender_display(),
            str(patient.get_age()),
            (patient.city or "-")[:15] + '...' if patient.city and len(patient.city) > 15 else (patient.city or "-"),
            "Activo" if patient.is_active else "Inactivo"
        ])
    
    # Create table
    table = Table(data, colWidths=[0.8*inch, 2.2*inch, 0.8*inch, 0.6*inch, 1.2*inch, 0.8*inch])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    if queryset.count() > 50:
        elements.append(Spacer(1, 10))
        note = Paragraph(
            f"Nota: Se muestran los primeros 50 pacientes de {queryset.count()} total.",
            styles['Normal']
        )
        elements.append(note)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_pacientes_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    
    return response

