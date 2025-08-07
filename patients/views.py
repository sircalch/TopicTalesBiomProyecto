from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import timedelta
import csv
import json
from io import BytesIO

from .models import Patient
from appointments.models import Appointment

@login_required
def patient_list(request):
    """
    List all patients with search and filtering capabilities
    """
    organization = request.user.profile.organization
    
    # Get all patients for this organization
    patients = Patient.objects.filter(organization=organization).order_by('-registration_date')
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        patients = patients.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(mother_last_name__icontains=search_query) |
            Q(patient_id__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    # Gender filter
    gender_filter = request.GET.get('gender')
    if gender_filter:
        patients = patients.filter(gender=gender_filter)
    
    # Age range filter
    age_range = request.GET.get('age_range')
    if age_range:
        today = timezone.now().date()
        if age_range == '0-17':
            birth_date_start = today - timedelta(days=17*365)
            patients = patients.filter(birth_date__gte=birth_date_start)
        elif age_range == '18-35':
            birth_date_start = today - timedelta(days=35*365)
            birth_date_end = today - timedelta(days=18*365)
            patients = patients.filter(birth_date__lte=birth_date_end, birth_date__gte=birth_date_start)
        elif age_range == '36-60':
            birth_date_start = today - timedelta(days=60*365)
            birth_date_end = today - timedelta(days=36*365)
            patients = patients.filter(birth_date__lte=birth_date_end, birth_date__gte=birth_date_start)
        elif age_range == '60+':
            birth_date_end = today - timedelta(days=60*365)
            patients = patients.filter(birth_date__lte=birth_date_end)
    
    # Blood type filter
    blood_type_filter = request.GET.get('blood_type')
    if blood_type_filter:
        patients = patients.filter(blood_type=blood_type_filter)
    
    # Statistics
    total_patients = Patient.objects.filter(organization=organization).count()
    active_patients = Patient.objects.filter(organization=organization, is_active=True).count()
    
    # New patients this month
    this_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    new_this_month = Patient.objects.filter(
        organization=organization,
        registration_date__gte=this_month_start
    ).count()
    
    # Patients with appointments today
    today = timezone.now().date()
    with_appointments_today = Patient.objects.filter(
        organization=organization,
        appointments__start_datetime__date=today
    ).distinct().count()
    
    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 25
    
    paginator = Paginator(patients, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Gestión de Pacientes',
        'patients': page_obj,
        'total_patients': total_patients,
        'active_patients': active_patients,
        'new_this_month': new_this_month,
        'with_appointments_today': with_appointments_today,
        'paginator': paginator,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        # Add search parameters to context
        'search_query': search_query,
        'gender_filter': gender_filter,
        'age_range': age_range,
        'blood_type_filter': blood_type_filter,
        'per_page': per_page,
    }
    
    return render(request, 'patients/list.html', context)

@login_required
def patient_create(request):
    """
    Create a new patient
    """
    from .forms import PatientForm
    
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES)
        if form.is_valid():
            patient = form.save(commit=False)
            patient.organization = request.user.profile.organization
            patient.created_by = request.user
            
            # Auto-generate patient_id if not provided
            if not patient.patient_id:
                organization = request.user.profile.organization
                last_patient = Patient.objects.filter(organization=organization).order_by('id').last()
                if last_patient:
                    try:
                        last_id = int(last_patient.patient_id.replace('PAC', ''))
                        patient.patient_id = f'PAC{str(last_id + 1).zfill(3)}'
                    except:
                        patient.patient_id = f'PAC{Patient.objects.filter(organization=organization).count() + 1:03d}'
                else:
                    patient.patient_id = 'PAC001'
            
            patient.save()
            messages.success(request, f'Paciente {patient.get_full_name()} creado exitosamente.')
            return redirect('patients:detail', patient_id=patient.id)
    else:
        form = PatientForm()
    
    return render(request, 'patients/create.html', {
        'title': 'Nuevo Paciente',
        'form': form
    })

@login_required
def patient_detail(request, patient_id):
    """
    Show patient details
    """
    patient = get_object_or_404(Patient, id=patient_id, organization=request.user.profile.organization)
    
    # Get recent appointments
    recent_appointments = patient.appointments.order_by('-start_datetime')[:5]
    
    # Get recent vital signs
    recent_vitals = patient.vital_signs.order_by('-recorded_at')[:5]
    
    # Get medical history
    try:
        medical_history = patient.medical_history
    except:
        medical_history = None
    
    # Get documents
    documents = patient.documents.order_by('-uploaded_at')[:10]
    
    context = {
        'title': f'Paciente: {patient.get_full_name()}',
        'patient': patient,
        'recent_appointments': recent_appointments,
        'recent_vitals': recent_vitals,
        'medical_history': medical_history,
        'documents': documents
    }
    
    return render(request, 'patients/detail.html', context)

@login_required
def patient_edit(request, patient_id):
    """
    Edit patient information
    """
    from .forms import PatientForm
    
    patient = get_object_or_404(Patient, id=patient_id, organization=request.user.profile.organization)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, f'Información de {patient.get_full_name()} actualizada exitosamente.')
            return redirect('patients:detail', patient_id=patient.id)
    else:
        form = PatientForm(instance=patient)
    
    return render(request, 'patients/edit.html', {
        'title': f'Editar: {patient.get_full_name()}',
        'form': form,
        'patient': patient
    })

@login_required
def medical_history(request, patient_id):
    """
    Manage patient medical history
    """
    from .forms import MedicalHistoryForm
    from .models import MedicalHistory
    
    patient = get_object_or_404(Patient, id=patient_id, organization=request.user.profile.organization)
    
    try:
        history = patient.medical_history
    except MedicalHistory.DoesNotExist:
        history = None
    
    if request.method == 'POST':
        if history:
            form = MedicalHistoryForm(request.POST, instance=history)
        else:
            form = MedicalHistoryForm(request.POST)
        
        if form.is_valid():
            medical_history = form.save(commit=False)
            medical_history.patient = patient
            medical_history.updated_by = request.user
            medical_history.save()
            messages.success(request, 'Historia médica actualizada exitosamente.')
            return redirect('patients:medical_history', patient_id=patient.id)
    else:
        form = MedicalHistoryForm(instance=history)
    
    return render(request, 'patients/medical_history.html', {
        'title': f'Historia Médica: {patient.get_full_name()}',
        'form': form,
        'patient': patient,
        'history': history
    })

@login_required
def vital_signs(request, patient_id):
    """
    Manage patient vital signs
    """
    from .forms import VitalSignsForm
    
    patient = get_object_or_404(Patient, id=patient_id, organization=request.user.profile.organization)
    
    if request.method == 'POST':
        form = VitalSignsForm(request.POST)
        if form.is_valid():
            vital_signs = form.save(commit=False)
            vital_signs.patient = patient
            vital_signs.recorded_by = request.user
            vital_signs.save()
            messages.success(request, 'Signos vitales registrados exitosamente.')
            return redirect('patients:vital_signs', patient_id=patient.id)
    else:
        form = VitalSignsForm()
    
    # Get all vital signs for this patient
    vital_signs_list = patient.vital_signs.order_by('-recorded_at')
    
    return render(request, 'patients/vital_signs.html', {
        'title': f'Signos Vitales: {patient.get_full_name()}',
        'form': form,
        'patient': patient,
        'vital_signs_list': vital_signs_list
    })

@login_required
def documents(request, patient_id):
    """
    Manage patient documents
    """
    from .forms import PatientDocumentForm
    
    patient = get_object_or_404(Patient, id=patient_id, organization=request.user.profile.organization)
    
    if request.method == 'POST':
        form = PatientDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.patient = patient
            document.uploaded_by = request.user
            document.save()
            messages.success(request, 'Documento subido exitosamente.')
            return redirect('patients:documents', patient_id=patient.id)
    else:
        form = PatientDocumentForm()
    
    # Get all documents for this patient
    documents_list = patient.documents.order_by('-uploaded_at')
    
    return render(request, 'patients/documents.html', {
        'title': f'Documentos: {patient.get_full_name()}',
        'form': form,
        'patient': patient,
        'documents_list': documents_list
    })


@login_required
def patient_search(request):
    """Búsqueda avanzada de pacientes"""
    organization = request.user.profile.organization
    patients = []
    
    if request.GET.get('search'):
        search_query = request.GET.get('search')
        patients = Patient.objects.filter(
            organization=organization
        ).filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(mother_last_name__icontains=search_query) |
            Q(patient_id__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        ).order_by('first_name', 'last_name')
    
    # Paginación
    paginator = Paginator(patients, 20)
    page_number = request.GET.get('page')
    patients_page = paginator.get_page(page_number)
    
    context = {
        'title': 'Buscar Pacientes',
        'patients': patients_page,
        'search_query': request.GET.get('search', ''),
    }
    
    return render(request, 'patients/search.html', context)


@login_required
def export_patients_csv(request):
    """
    Export patients to CSV format
    """
    organization = request.user.profile.organization
    patients = Patient.objects.filter(organization=organization).order_by('first_name', 'last_name')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="pacientes_{organization.name}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID Paciente', 'Nombre', 'Apellidos', 'Email', 'Telefono', 
        'Fecha Nacimiento', 'Genero', 'Fecha Registro', 'Estado'
    ])
    
    for patient in patients:
        writer.writerow([
            patient.patient_id,
            patient.first_name,
            f"{patient.last_name} {patient.mother_last_name}".strip(),
            patient.email or '',
            patient.phone_number or '',
            patient.birth_date.strftime('%Y-%m-%d') if patient.birth_date else '',
            patient.get_gender_display(),
            patient.registration_date.strftime('%Y-%m-%d'),
            'Activo' if patient.is_active else 'Inactivo'
        ])
    
    return response


@login_required
def export_patients_excel(request):
    """
    Export patients to Excel format (using CSV for now, can be enhanced with openpyxl)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        organization = request.user.profile.organization
        patients = Patient.objects.filter(organization=organization).order_by('first_name', 'last_name')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacientes"
        
        # Headers
        headers = [
            'ID Paciente', 'Nombre', 'Apellidos', 'Email', 'Teléfono', 
            'Fecha Nacimiento', 'Género', 'Fecha Registro', 'Estado'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data
        for row_num, patient in enumerate(patients, 2):
            ws.cell(row=row_num, column=1, value=patient.patient_id)
            ws.cell(row=row_num, column=2, value=patient.first_name)
            ws.cell(row=row_num, column=3, value=f"{patient.last_name} {patient.mother_last_name}".strip())
            ws.cell(row=row_num, column=4, value=patient.email or '')
            ws.cell(row=row_num, column=5, value=patient.phone_number or '')
            ws.cell(row=row_num, column=6, value=patient.birth_date.strftime('%Y-%m-%d') if patient.birth_date else '')
            ws.cell(row=row_num, column=7, value=patient.get_gender_display())
            ws.cell(row=row_num, column=8, value=patient.registration_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=9, value='Activo' if patient.is_active else 'Inactivo')
        
        # Adjust column widths
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="pacientes_{organization.name}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return export_patients_csv(request)


@login_required
def export_patients_pdf(request):
    """
    Export patients to PDF format (basic implementation)
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        organization = request.user.profile.organization
        patients = Patient.objects.filter(organization=organization).order_by('first_name', 'last_name')
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
        )
        
        # Content
        story = []
        
        # Title
        title = Paragraph(f"Lista de Pacientes - {organization.name}", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Table data
        data = [['ID', 'Nombre Completo', 'Email', 'Teléfono', 'Fecha Registro']]
        
        for patient in patients:
            data.append([
                patient.patient_id,
                f"{patient.first_name} {patient.last_name} {patient.mother_last_name}".strip(),
                patient.email or '',
                patient.phone_number or '',
                patient.registration_date.strftime('%d/%m/%Y')
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pacientes_{organization.name}_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        return response
        
    except ImportError:
        # Fallback message if reportlab is not available
        return HttpResponse(
            "PDF export requires reportlab library. Please contact administrator.",
            content_type='text/plain'
        )


@login_required
def patient_quick_actions(request):
    """
    Handle quick actions for patients via AJAX
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            patient_id = data.get('patient_id')
            
            patient = get_object_or_404(Patient, id=patient_id, organization=request.user.profile.organization)
            
            if action == 'toggle_active':
                patient.is_active = not patient.is_active
                patient.save()
                return JsonResponse({
                    'success': True,
                    'message': f'Paciente {"activado" if patient.is_active else "desactivado"} exitosamente',
                    'new_status': patient.is_active
                })
            
            elif action == 'quick_appointment':
                # Redirect to appointment creation with patient pre-selected
                return JsonResponse({
                    'success': True,
                    'redirect': f'/appointments/create/?patient={patient_id}'
                })
            
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Acción no válida'
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})
